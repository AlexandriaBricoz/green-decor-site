import json
import logging

from django.conf import settings
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from .forms import ContactForm
from .models import ContactRequest
from .telegram import (
    answer_callback_query,
    edit_message_reply_markup,
    is_configured,
    send_document,
    send_message,
)

logger = logging.getLogger(__name__)


def _client_ip(request):
    xff = request.META.get("HTTP_X_FORWARDED_FOR", "")
    if xff:
        return xff.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


def submit_contact(request):
    if request.method != "POST":
        return redirect("catalog:home")

    form = ContactForm(request.POST)

    if form.is_spam():
        # Тихо принимаем заявку от бота, ничего не сохраняем.
        messages.success(
            request,
            "Заявка отправлена. Мы свяжемся с вами в ближайшее время.",
        )
        return redirect("catalog:home")

    if form.is_valid():
        req = form.save(commit=False)
        req.consent_given = True
        req.consent_at = timezone.now()
        req.source_ip = _client_ip(request)
        req.user_agent = request.META.get("HTTP_USER_AGENT", "")[:500]
        req.save()
        messages.success(
            request,
            "Заявка отправлена. Мы свяжемся с вами в ближайшее время.",
        )
        return redirect("catalog:home")

    messages.error(request, "Проверьте правильность заполнения формы.")
    return render(
        request,
        "catalog/home.html",
        {"contact_form": form, "scroll_to": "contact"},
    )


HELP_TEMPLATE = (
    "<b>Бот {brand}</b>\n"
    "Команды в этом чате:\n"
    "/table — Excel со всеми заявками\n"
    "/start — переустановить чат как получатель уведомлений\n"
    "/help — эта справка"
)


def _help_text() -> str:
    from catalog.models import SiteSettings
    return HELP_TEMPLATE.format(brand=SiteSettings.load().brand_name or "Green Decor")


def _requests_xlsx_bytes() -> bytes:
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    header = ["Дата", "Имя", "Контакт", "Компания", "Тема", "Сообщение", "IP", "Обработано"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="006C0C")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26

    for r in ContactRequest.objects.all().order_by("-created_at"):
        ws.append([
            timezone.localtime(r.created_at).strftime("%d.%m.%Y %H:%M"),
            r.name,
            r.contact,
            r.company,
            r.get_subject_display(),
            r.message,
            r.source_ip or "",
            "да" if r.processed else "нет",
        ])

    widths = [16, 22, 26, 26, 22, 44, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _save_chat_id(chat_id: str) -> None:
    from catalog.models import SiteSettings
    obj = SiteSettings.load()
    obj.telegram_chat_id = str(chat_id)
    obj.save(update_fields=["telegram_chat_id"])


def _current_chat_id() -> str:
    from catalog.models import SiteSettings
    return SiteSettings.load().telegram_chat_id or ""


def _handle_message(msg: dict) -> None:
    text = (msg.get("text") or "").strip()
    chat = msg.get("chat") or {}
    chat_id = chat.get("id")
    if chat_id is None:
        return

    # /start — привязка чата (можно писать в любом чате, чтобы привязать)
    if text.startswith("/start"):
        _save_chat_id(chat_id)
        send_message(
            chat_id,
            "✅ Бот привязан к этому чату.\n\n" + _help_text(),
        )
        return

    # Дальше — только для привязанного чата
    if str(chat_id) != _current_chat_id():
        return

    if text.startswith("/table"):
        try:
            data = _requests_xlsx_bytes()
        except Exception:
            logger.exception("не удалось собрать xlsx")
            send_message(chat_id, "⚠ Не удалось собрать таблицу. Логи на сервере.")
            return
        total = ContactRequest.objects.count()
        new = ContactRequest.objects.filter(processed=False).count()
        filename = f"greendecor-requests-{timezone.localdate():%Y-%m-%d}.xlsx"
        caption = f"Всего заявок: {total} · Не обработано: {new}"
        send_document(chat_id, filename, data, caption=caption)
        return

    if text.startswith("/help"):
        send_message(chat_id, _help_text())
        return


def _handle_callback(cq: dict) -> None:
    data = cq.get("data") or ""
    msg = cq.get("message") or {}
    chat_id = (msg.get("chat") or {}).get("id")
    cq_id = cq.get("id")
    if chat_id is None or cq_id is None:
        return
    if str(chat_id) != _current_chat_id():
        answer_callback_query(cq_id, "Этот чат не привязан к боту.")
        return
    if data.startswith("processed:"):
        try:
            pk = int(data.split(":", 1)[1])
        except (ValueError, IndexError):
            answer_callback_query(cq_id, "Некорректные данные.")
            return
        obj = ContactRequest.objects.filter(pk=pk).first()
        if not obj:
            answer_callback_query(cq_id, "Заявка не найдена.")
            return
        if not obj.processed:
            obj.processed = True
            obj.save(update_fields=["processed"])
        answer_callback_query(cq_id, "Отмечено как обработано.")
        edit_message_reply_markup(chat_id, msg.get("message_id"))
        return
    answer_callback_query(cq_id, "Неизвестное действие.")


@csrf_exempt
@require_POST
def telegram_webhook(request, secret: str):
    expected = getattr(settings, "TELEGRAM_WEBHOOK_SECRET", "")
    if not expected or secret != expected:
        return HttpResponse(status=403)
    header_secret = request.headers.get("X-Telegram-Bot-Api-Secret-Token", "")
    if header_secret != expected:
        return HttpResponse(status=403)
    if not is_configured():
        return HttpResponse(status=503)
    try:
        update = json.loads(request.body.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return HttpResponse(status=400)
    try:
        if "callback_query" in update:
            _handle_callback(update["callback_query"])
        elif "message" in update:
            _handle_message(update["message"])
    except Exception:
        logger.exception("необработанная ошибка в telegram_webhook")
    return JsonResponse({"ok": True})
