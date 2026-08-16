from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import models
from django.db.models import F, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from contacts.forms import ContactForm
from contacts.models import ContactRequest

from .forms import PlantForm, SiteSettingsForm
from .models import Plant, SiteSettings


def home(request):
    contact_form = ContactForm()
    return render(request, "catalog/home.html", {"contact_form": contact_form})


def privacy(request):
    return render(request, "catalog/privacy.html")


def terms(request):
    return render(request, "catalog/terms.html")


@login_required
def inventory_list(request):
    base_qs = Plant.objects.all()
    search = request.GET.get("q", "").strip()
    if search:
        base_qs = base_qs.filter(name__icontains=search)

    visibility = request.GET.get("visibility", "all")
    if visibility == "public":
        qs = base_qs.filter(is_public=True)
    elif visibility == "hidden":
        qs = base_qs.filter(is_public=False)
    else:
        visibility = "all"
        qs = base_qs

    stats = qs.aggregate(
        total_qty=Sum("quantity"),
        total_value=Sum(F("quantity") * F("price")),
    )
    low_stock_count = qs.filter(quantity__lte=Plant.LOW_STOCK_THRESHOLD).count()
    public_count = base_qs.filter(is_public=True).count()
    hidden_count = base_qs.filter(is_public=False).count()
    total_count = base_qs.count()

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "catalog/inventory_list.html",
        {
            "page_obj": page_obj,
            "paginator": paginator,
            "total_qty": stats["total_qty"] or 0,
            "total_value": stats["total_value"] or Decimal("0"),
            "low_stock_count": low_stock_count,
            "public_count": public_count,
            "hidden_count": hidden_count,
            "total_count": total_count,
            "search": search,
            "visibility": visibility,
        },
    )


@login_required
def plant_create(request):
    if request.method == "POST":
        form = PlantForm(request.POST, request.FILES)
        if form.is_valid():
            plant = form.save()
            messages.success(request, f"Растение «{plant.name}» добавлено в каталог.")
            return redirect("catalog:inventory")
    else:
        form = PlantForm()
    return render(
        request,
        "catalog/plant_form.html",
        {"form": form, "is_edit": False},
    )


@login_required
def plant_edit(request, pk):
    plant = get_object_or_404(Plant, pk=pk)
    if request.method == "POST":
        form = PlantForm(request.POST, request.FILES, instance=plant)
        if form.is_valid():
            form.save()
            messages.success(request, f"Растение «{plant.name}» обновлено.")
            return redirect("catalog:inventory")
    else:
        form = PlantForm(instance=plant)
    return render(
        request,
        "catalog/plant_form.html",
        {"form": form, "plant": plant, "is_edit": True},
    )


@login_required
@require_POST
def plant_toggle_visibility(request, pk):
    plant = get_object_or_404(Plant, pk=pk)
    plant.is_public = not plant.is_public
    plant.save(update_fields=["is_public", "updated_at"])

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"id": plant.pk, "is_public": plant.is_public})

    state = "опубликовано" if plant.is_public else "скрыто"
    messages.success(request, f"Растение «{plant.name}» {state} в публичном прайсе.")
    return redirect("catalog:inventory")


@login_required
def plant_delete(request, pk):
    plant = get_object_or_404(Plant, pk=pk)
    if request.method == "POST":
        name = plant.name
        plant.delete()
        messages.success(request, f"Растение «{name}» удалено.")
        return redirect("catalog:inventory")
    return render(request, "catalog/plant_delete.html", {"plant": plant})


@login_required
def site_settings_edit(request):
    obj = SiteSettings.load()
    if request.method == "POST":
        form = SiteSettingsForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Настройки сайта сохранены.")
            return redirect("catalog:settings")
    else:
        form = SiteSettingsForm(instance=obj)
    return render(
        request,
        "catalog/settings_form.html",
        {"form": form, "obj": obj},
    )


@login_required
def inbox_list(request):
    base_qs = ContactRequest.objects.all()

    search = request.GET.get("q", "").strip()
    if search:
        base_qs = base_qs.filter(
            models.Q(name__icontains=search)
            | models.Q(contact__icontains=search)
            | models.Q(company__icontains=search)
            | models.Q(message__icontains=search)
        )

    status = request.GET.get("status", "all")
    if status == "new":
        qs = base_qs.filter(processed=False)
    elif status == "done":
        qs = base_qs.filter(processed=True)
    else:
        status = "all"
        qs = base_qs

    total_count = base_qs.count()
    new_count = base_qs.filter(processed=False).count()
    done_count = total_count - new_count

    paginator = Paginator(qs.order_by("-created_at"), 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "catalog/inbox_list.html",
        {
            "page_obj": page_obj,
            "paginator": paginator,
            "total_count": total_count,
            "new_count": new_count,
            "done_count": done_count,
            "search": search,
            "status": status,
        },
    )


@login_required
def inbox_detail(request, pk):
    obj = get_object_or_404(ContactRequest, pk=pk)
    return render(request, "catalog/inbox_detail.html", {"obj": obj})


@login_required
@require_POST
def inbox_toggle(request, pk):
    obj = get_object_or_404(ContactRequest, pk=pk)
    obj.processed = not obj.processed
    obj.save(update_fields=["processed"])
    state = "обработана" if obj.processed else "возвращена в новые"
    messages.success(request, f"Заявка от {obj.name} {state}.")
    next_url = request.POST.get("next") or reverse("catalog:inbox")
    return redirect(next_url)


@login_required
def inbox_delete(request, pk):
    obj = get_object_or_404(ContactRequest, pk=pk)
    if request.method == "POST":
        name = obj.name
        obj.delete()
        messages.success(request, f"Заявка от {name} удалена.")
        return redirect("catalog:inbox")
    return render(request, "catalog/inbox_delete.html", {"obj": obj})


@login_required
def inbox_export_xlsx(request):
    from contacts.views import _requests_xlsx_bytes

    data = _requests_xlsx_bytes()
    resp = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    today = timezone.localdate().strftime("%Y-%m-%d")
    resp["Content-Disposition"] = f'attachment; filename="greendecor-requests-{today}.xlsx"'
    return resp


def _plant_thumbnail_bytes(plant, size=(60, 60)):
    """Return a BytesIO with a thumbnail JPEG of plant.photo, or None."""
    if not plant.photo:
        return None
    try:
        from io import BytesIO

        from PIL import Image as PILImage
    except ImportError:
        return None
    try:
        with PILImage.open(plant.photo.path) as im:
            im = im.convert("RGB")
            im.thumbnail(size, PILImage.LANCZOS)
            buf = BytesIO()
            im.save(buf, format="JPEG", quality=82, optimize=True)
            buf.seek(0)
            return buf
    except (FileNotFoundError, OSError):
        return None


def price_list_xlsx_legacy(request):
    return redirect("catalog:price_xlsx", permanent=True)


def price_list_xlsx(request):
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            "openpyxl не установлен. Выполните: pip install openpyxl",
            status=500,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"

    header = [
        "Фото",
        "Артикул",
        "Наименование",
        "Возраст (лет)",
        "Размер (см)",
        "Количество",
        "Цена, ₽/шт",
        "Медиа",
    ]
    ws.append(header)

    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill = PatternFill("solid", fgColor="006C0C")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[1].height = 30

    body_alignment = Alignment(vertical="center")
    image_buffers = []  # keep references alive until wb.save()

    for row_idx, plant in enumerate(
        Plant.objects.filter(is_public=True).order_by("name"), start=2
    ):
        ws.cell(row=row_idx, column=1, value="")
        ws.cell(row=row_idx, column=2, value=plant.sku).alignment = body_alignment
        ws.cell(row=row_idx, column=3, value=plant.name).alignment = body_alignment
        ws.cell(row=row_idx, column=4, value=plant.age_years).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=row_idx, column=5, value=plant.size_cm).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=row_idx, column=6, value=plant.quantity).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        price_cell = ws.cell(row=row_idx, column=7, value=float(plant.price))
        price_cell.number_format = "#,##0.00 ₽"
        price_cell.alignment = Alignment(horizontal="right", vertical="center")

        media_cell = ws.cell(row=row_idx, column=8)
        media_cell.alignment = Alignment(horizontal="center", vertical="center")
        if plant.external_url:
            media_cell.value = "Смотреть"
            media_cell.hyperlink = plant.external_url
            media_cell.font = Font(color="0563C1", underline="single")

        ws.row_dimensions[row_idx].height = 48

        buf = _plant_thumbnail_bytes(plant)
        if buf is not None:
            image_buffers.append(buf)
            try:
                xl_img = XLImage(buf)
                xl_img.width = 58
                xl_img.height = 58
                ws.add_image(xl_img, f"A{row_idx}")
            except Exception:
                pass

    widths = [10, 18, 38, 14, 14, 14, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    today = timezone.now().strftime("%Y-%m-%d")
    response["Content-Disposition"] = f'attachment; filename="green-decor-price-{today}.xlsx"'
    wb.save(response)
    return response
